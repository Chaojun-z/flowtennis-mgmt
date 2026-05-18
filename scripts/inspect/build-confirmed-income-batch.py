#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


def norm_text(value: object) -> str:
    return (
        str(value or "")
        .replace("，", ",")
        .replace("、", ",")
        .replace("／", "/")
        .replace("课时", "")
        .replace("扣一课时", "扣1课时")
        .replace("扣一节", "扣1课时")
        .replace("扣一次", "扣1次")
        .strip()
    )


def money(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def extract_first_amount(text: str) -> float | None:
    matched = re.search(r"(\d+(?:\.\d+)?)", text or "")
    return float(matched.group(1)) if matched else None


def classify(row: dict[str, object]) -> dict[str, object]:
    line_no = str(row.get("原表行号") or "")
    answer = norm_text(row.get("你的回答"))
    customer = str(row.get("客户") or "")
    original_pay = str(row.get("支付方式") or "")
    supplemental_note = str(row.get("补充备注") or "")
    due = money(row.get("应收"))
    paid = money(row.get("实收"))

    result = {
        "status": "ready",
        "parseType": "",
        "businessType": "",
        "campusId": "mabao",
        "campusName": "顺义马坡",
        "accountingScope": "mabao_only",
        "paymentChannel": "",
        "recognizeRevenue": "",
        "cashDelta": "",
        "recognizedRevenueDelta": "",
        "deferredRevenueDelta": "",
        "studentName": "",
        "packageName": "",
        "lessonCount": "",
        "couponAmount": "",
        "wechatTopupAmount": "",
        "note": "",
    }

    # 人工已在对话中明确确认的覆盖规则，优先级最高
    if line_no == "239":
        result.update(
            {
                "parseType": "booking_income",
                "businessType": "场地费收入",
                "paymentChannel": "微信",
                "recognizeRevenue": "是",
                "cashDelta": 660,
                "recognizedRevenueDelta": 660,
                "deferredRevenueDelta": 0,
                "note": "人工确认：朝珺自收课费，场地费 660 转给俱乐部",
            }
        )
        return result
    if line_no == "40":
        result.update(
            {
                "parseType": "free_gift",
                "businessType": "免费赠送",
                "paymentChannel": "无",
                "recognizeRevenue": "否",
                "cashDelta": 0,
                "recognizedRevenueDelta": 0,
                "deferredRevenueDelta": 0,
                "note": "人工确认：收费 0 + 免费赠送 220，按免费赠送处理",
            }
        )
        return result
    if line_no == "100":
        result["parseType"] = "booking_income"
        result["businessType"] = "订场收入"
        result["paymentChannel"] = "微信"
        result["recognizeRevenue"] = "是"
        result["cashDelta"] = 220
        result["recognizedRevenueDelta"] = 220
        result["deferredRevenueDelta"] = 0
        result["note"] = "人工确认：订场2h，收费220 + 免费赠送 220"
        return result
    if line_no == "125":
        result["parseType"] = "booking_income"
        result["businessType"] = "订场收入"
        result["paymentChannel"] = "微信"
        result["recognizeRevenue"] = "是"
        result["cashDelta"] = 220
        result["recognizedRevenueDelta"] = 220
        result["deferredRevenueDelta"] = 0
        result["note"] = "人工确认：订场2h，收费220 + 免费赠送 220"
        return result
    if line_no in ("127", "154"):
        result.update(
            {
                "parseType": "dp_coupon",
                "businessType": "大众点评私教券",
                "paymentChannel": "大众点评",
                "recognizeRevenue": "是",
                "cashDelta": 320,
                "recognizedRevenueDelta": 320,
                "deferredRevenueDelta": 0,
                "couponAmount": 320,
                "wechatTopupAmount": 0,
                "note": "人工确认：私教券，应收实收均 320",
            }
        )
        return result
    if line_no == "439":
        result.update(
            {
                "parseType": "booking_income",
                "businessType": "场地费收入",
                "paymentChannel": "微信",
                "recognizeRevenue": "是",
                "cashDelta": 140,
                "recognizedRevenueDelta": 140,
                "deferredRevenueDelta": 0,
                "note": "人工确认：场地费 140，微信收款",
            }
        )
        return result
    if line_no == "540":
        result.update(
            {
                "parseType": "package_consume",
                "businessType": "课包划扣",
                "paymentChannel": "课包划扣",
                "recognizeRevenue": "是",
                "cashDelta": 0,
                "recognizedRevenueDelta": 1200,
                "deferredRevenueDelta": -1200,
                "studentName": "oliver",
                "packageName": "待按现有课包匹配",
                "lessonCount": 2,
                "note": "人工确认：课包划扣2课时，共1200（1课时600）",
            }
        )
        return result
    if line_no == "516":
        result.update(
            {
                "parseType": "cross_campus_consume_trace",
                "businessType": "跨校区课包留痕",
                "campusId": "shilipu",
                "campusName": "朝阳十里堡",
                "accountingScope": "external_trace_only",
                "paymentChannel": "课包划扣",
                "recognizeRevenue": "否",
                "cashDelta": 0,
                "recognizedRevenueDelta": 0,
                "deferredRevenueDelta": 0,
                "studentName": "陈川",
                "packageName": "60课时课包",
                "lessonCount": 2,
                "note": "人工确认：十里堡课包扣2课时，仅留痕，不计入马坡收入",
            }
        )
        return result
    if line_no == "776":
        recognized = due or 0
        deferred = -(recognized - 80)
        result.update(
            {
                "parseType": "package_consume",
                "businessType": "课包划扣+补差价",
                "paymentChannel": "课包划扣+微信补差",
                "recognizeRevenue": "是",
                "cashDelta": 80,
                "recognizedRevenueDelta": recognized,
                "deferredRevenueDelta": deferred,
                "studentName": customer.split()[0] if customer else "",
                "packageName": "待按现有课包匹配",
                "lessonCount": 1,
                "note": "人工确认：课包扣1次，另收80微信",
            }
        )
        return result
    if line_no == "107":
        result.update(
            {
                "parseType": "package_consume",
                "businessType": "课包划扣",
                "paymentChannel": "课包划扣",
                "recognizeRevenue": "是",
                "cashDelta": 0,
                "recognizedRevenueDelta": due or paid,
                "deferredRevenueDelta": -(due or paid),
                "studentName": "高珺",
                "packageName": "成人1v1 10课时非黄金时间课包",
                "lessonCount": 1,
                "note": "人工确认：课包扣1次；备注：不知道丹丹的课包多少钱",
            }
        )
        return result
    if line_no == "364":
        result.update(
            {
                "parseType": "package_consume",
                "businessType": "课包划扣",
                "paymentChannel": "课包划扣",
                "recognizeRevenue": "是",
                "cashDelta": 0,
                "recognizedRevenueDelta": due or paid,
                "deferredRevenueDelta": -(due or paid),
                "studentName": "高珺",
                "packageName": "1v1 10课时非黄金时间课包",
                "lessonCount": 1,
                "note": "人工确认：高珺/1v1 10课时非黄金时间课包，扣1课时；备注：不知道丹丹的课包多少钱",
            }
        )
        return result
    if line_no == "951":
        amount = 440
        result.update(
            {
                "parseType": "course_income",
                "businessType": "课程收入",
                "paymentChannel": "微信",
                "recognizeRevenue": "是",
                "cashDelta": amount,
                "recognizedRevenueDelta": amount,
                "deferredRevenueDelta": 0,
                "note": "人工确认：马坡已收440微信",
            }
        )
        return result
    if line_no in ("1035", "1138", "1144", "1352", "1356"):
        result.update(
            {
                "parseType": "camp_income",
                "businessType": "冬训营收入",
                "paymentChannel": "沿用原表",
                "recognizeRevenue": "是",
                "cashDelta": 280,
                "recognizedRevenueDelta": 280,
                "deferredRevenueDelta": 0,
                "note": "人工确认：冬训营按每天2小时固定280导入",
            }
        )
        return result
    if line_no == "1167":
        amount = paid or due
        result.update(
            {
                "parseType": "course_income",
                "businessType": "课程收入",
                "paymentChannel": "微信",
                "recognizeRevenue": "是",
                "cashDelta": amount,
                "recognizedRevenueDelta": amount,
                "deferredRevenueDelta": 0,
                "note": "人工确认：课程/微信",
            }
        )
        return result
    if line_no == "1168":
        amount = 480
        result.update(
            {
                "parseType": "package_consume",
                "businessType": "课包划扣",
                "paymentChannel": "课包划扣",
                "recognizeRevenue": "是",
                "cashDelta": 0,
                "recognizedRevenueDelta": amount,
                "deferredRevenueDelta": -amount,
                "studentName": "丫丫",
                "packageName": "1v1 10课时非黄金时间课包",
                "lessonCount": 1,
                "note": "人工确认：课包划扣1课时 480",
            }
        )
        return result
    if line_no == "212":
        result["status"] = "split"
        result["parseType"] = "multi_package_split"
        result["businessType"] = "课包划扣"
        result["paymentChannel"] = "课包划扣"
        result["recognizeRevenue"] = "是"
        result["note"] = "人工确认：拆成 misha 1次 600 + 黄总 1次 600"
        return result

    if not answer:
        if any(token in customer for token in ["装修", "洗场地", "清洗场地"]):
            result.update(
                {
                    "parseType": "internal_use",
                    "businessType": "内部占用",
                    "paymentChannel": "无",
                    "recognizeRevenue": "否",
                    "cashDelta": 0,
                    "recognizedRevenueDelta": 0,
                    "deferredRevenueDelta": 0,
                    "note": "按内部占用默认处理",
                }
            )
            return result
        result["status"] = "missing"
        result["note"] = "你的回答为空"
        return result

    # 最后一列是运营后补的说明，优先级高于简化答案
    if "应收和实收都是660" in supplemental_note:
        result.update(
            {
                "parseType": "booking_income",
                "businessType": "场地费收入",
                "paymentChannel": "微信",
                "recognizeRevenue": "是",
                "cashDelta": 660,
                "recognizedRevenueDelta": 660,
                "deferredRevenueDelta": 0,
                "note": "按补充备注修正：朝珺自收课费，场地费 660 转给俱乐部",
            }
        )
        return result

    if any(token in supplemental_note for token in ["需要和丹丹姐确认", "需要问丹丹姐", "需要何丹丹姐核对", "不知道多钱", "不知道怎么支付"]):
        result["status"] = "pending"
        result["note"] = supplemental_note
        return result

    if "十里堡的学员" in supplemental_note:
        result["status"] = "ready"
        result["parseType"] = "cross_campus_consume_trace"
        result["businessType"] = "跨校区课包留痕"
        result["campusId"] = "shilipu"
        result["campusName"] = "朝阳十里堡"
        result["accountingScope"] = "external_trace_only"
        result["paymentChannel"] = "课包划扣"
        result["recognizeRevenue"] = "否"
        result["cashDelta"] = 0
        result["recognizedRevenueDelta"] = 0
        result["deferredRevenueDelta"] = 0
        result["note"] = supplemental_note
        return result

    if "需要补80差价" in supplemental_note:
        recognized = due or paid
        result["status"] = "ready"
        result["parseType"] = "package_consume"
        result["businessType"] = "课包划扣+补差价"
        result["paymentChannel"] = "课包划扣+微信补差"
        result["recognizeRevenue"] = "是"
        result["cashDelta"] = 80
        result["recognizedRevenueDelta"] = recognized
        result["deferredRevenueDelta"] = -(recognized - 80)
        result["lessonCount"] = 1
        result["note"] = supplemental_note
        return result

    if "没有摊到每一天" in supplemental_note or "每天2小时的费用应该是4个课包的总价除5" in supplemental_note:
        result["status"] = "ready"
        result["parseType"] = "camp_income"
        result["businessType"] = "冬训营收入"
        result["paymentChannel"] = "沿用原表"
        result["recognizeRevenue"] = "是"
        result["cashDelta"] = 280
        result["recognizedRevenueDelta"] = 280
        result["deferredRevenueDelta"] = 0
        result["note"] = supplemental_note
        return result

    if "分别有自己的课包" in supplemental_note:
        result["status"] = "split"
        result["parseType"] = "multi_package_split"
        result["businessType"] = "课包划扣"
        result["paymentChannel"] = "课包划扣"
        result["recognizeRevenue"] = "是"
        result["note"] = supplemental_note
        return result

    if answer in ("内部占用", "免费赠送", "免费课"):
        mapping = {
            "内部占用": ("internal_use", "内部占用"),
            "免费赠送": ("free_gift", "免费赠送"),
            "免费课": ("free_course", "免费课"),
        }
        parse_type, business_type = mapping[answer]
        result.update(
            {
                "parseType": parse_type,
                "businessType": business_type,
                "paymentChannel": "无",
                "recognizeRevenue": "否",
                "cashDelta": 0,
                "recognizedRevenueDelta": 0,
                "deferredRevenueDelta": 0,
                "note": "运营已确认",
            }
        )
        return result

    if "训练营收入" in answer:
        if "冬训营" in customer:
            result["status"] = "ready"
            result["parseType"] = "camp_income"
            result["businessType"] = "冬训营收入"
            result["paymentChannel"] = "沿用原表"
            result["recognizeRevenue"] = "是"
            result["cashDelta"] = 280
            result["recognizedRevenueDelta"] = 280
            result["deferredRevenueDelta"] = 0
            result["note"] = "人工确认：冬训营按每天2小时固定280导入"
            return result
        result.update(
            {
                "parseType": "camp_income",
                "businessType": "训练营收入",
                "paymentChannel": "沿用原表",
                "recognizeRevenue": "是",
                "cashDelta": paid,
                "recognizedRevenueDelta": paid,
                "deferredRevenueDelta": 0,
                "note": "可直接导入训练营收入",
            }
        )
        return result

    if "忘记取消场地" in answer:
        result.update(
            {
                "parseType": "internal_use",
                "businessType": "内部占用",
                "paymentChannel": "无",
                "recognizeRevenue": "否",
                "cashDelta": 0,
                "recognizedRevenueDelta": 0,
                "deferredRevenueDelta": 0,
                "note": "按内部占用/取消漏记处理",
            }
        )
        return result

    if answer in ("专项课,微信收款", "专项课, 微信收款", "专项课，微信收款"):
        result.update(
            {
                "parseType": "course_income",
                "businessType": "专项课",
                "paymentChannel": "微信",
                "recognizeRevenue": "是",
                "cashDelta": paid or due,
                "recognizedRevenueDelta": paid or due,
                "deferredRevenueDelta": 0,
                "note": "专项课微信收款",
            }
        )
        return result

    if answer.startswith("课程收入") or "微信已收" in answer:
        if "大众点评" in answer:
            if paid and due and paid == due:
                result.update(
                    {
                        "parseType": "dp_coupon",
                        "businessType": "大众点评券",
                        "paymentChannel": "大众点评",
                        "recognizeRevenue": "是",
                        "cashDelta": paid,
                        "recognizedRevenueDelta": paid,
                        "deferredRevenueDelta": 0,
                        "couponAmount": paid,
                        "wechatTopupAmount": 0,
                        "note": "按大众点评券全额处理",
                    }
                )
                return result
            result["status"] = "pending"
            result["parseType"] = "course_dp"
            result["businessType"] = "课程收入"
            result["paymentChannel"] = "大众点评"
            result["recognizeRevenue"] = "是"
            result["cashDelta"] = paid
            result["recognizedRevenueDelta"] = paid
            result["deferredRevenueDelta"] = 0
            result["note"] = "课程来自大众点评，但没写券金额"
            return result
        result.update(
            {
                "parseType": "course_income",
                "businessType": "课程收入",
                "paymentChannel": "微信" if "微信" in answer else "沿用原表",
                "recognizeRevenue": "是",
                "cashDelta": extract_first_amount(answer) or paid or due,
                "recognizedRevenueDelta": extract_first_amount(answer) or paid or due,
                "deferredRevenueDelta": 0,
                "note": "课程收入已明确",
            }
        )
        return result

    if ("订场收入" in answer or answer.startswith("订场/")) and (
        "已收微信" in answer or "微信收款" in answer or "微信收" in answer
    ):
        if "免费赠送" in answer and not (paid and due == paid):
            if paid == 0:
                result.update(
                    {
                        "parseType": "free_gift",
                        "businessType": "免费赠送",
                        "paymentChannel": "无",
                        "recognizeRevenue": "否",
                        "cashDelta": 0,
                        "recognizedRevenueDelta": 0,
                        "deferredRevenueDelta": 0,
                        "note": f"场地免费赠送 {max(due - paid, 0):.2f}",
                    }
                )
                return result
            result["status"] = "split"
            result["parseType"] = "booking_plus_free"
            result["businessType"] = "订场收入+免费赠送"
            result["paymentChannel"] = "微信"
            result["recognizeRevenue"] = "部分确认"
            result["cashDelta"] = paid
            result["recognizedRevenueDelta"] = paid
            result["deferredRevenueDelta"] = 0
            result["note"] = f"需要拆分成收费 {paid:.2f} + 免费赠送 {max(due - paid, 0):.2f}"
            return result
        result.update(
            {
                "parseType": "booking_income",
                "businessType": "订场收入",
                "paymentChannel": "微信",
                "recognizeRevenue": "是",
                "cashDelta": paid or extract_first_amount(answer) or due,
                "recognizedRevenueDelta": paid or extract_first_amount(answer) or due,
                "deferredRevenueDelta": 0,
                "note": "订场收入已明确",
            }
        )
        return result

    if answer == "订场收入 / 免费赠送" or answer == "订场收入/免费赠送":
        if paid > 0 and due == paid:
            result.update(
                {
                    "parseType": "booking_income",
                    "businessType": "订场收入",
                    "paymentChannel": original_pay or "沿用原表",
                    "recognizeRevenue": "是",
                    "cashDelta": paid,
                    "recognizedRevenueDelta": paid,
                    "deferredRevenueDelta": 0,
                    "note": "原表应收=实收，按订场收入导入，保留备注里的赠送信息",
                }
            )
            return result
        result["status"] = "split"
        result["parseType"] = "booking_plus_free"
        result["businessType"] = "订场收入+免费赠送"
        result["paymentChannel"] = original_pay or "微信"
        result["recognizeRevenue"] = "部分确认"
        result["cashDelta"] = paid
        result["recognizedRevenueDelta"] = paid
        result["deferredRevenueDelta"] = 0
        result["note"] = f"需要拆分成收费 {paid:.2f} + 免费赠送 {max(due - paid, 0):.2f}"
        return result

    if "1小时订场收入/1小时免费赠送" in answer:
        result["status"] = "split"
        result["parseType"] = "booking_plus_free"
        result["businessType"] = "订场收入+免费赠送"
        result["paymentChannel"] = "微信"
        result["recognizeRevenue"] = "部分确认"
        result["cashDelta"] = paid
        result["recognizedRevenueDelta"] = paid
        result["deferredRevenueDelta"] = 0
        result["note"] = f"需要拆分成收费 {paid:.2f} + 免费赠送 {max(due - paid, 0):.2f}"
        return result

    if "已收小程序" in answer:
        result.update(
            {
                "parseType": "booking_income",
                "businessType": "订场收入",
                "paymentChannel": "小程序",
                "recognizeRevenue": "是",
                "cashDelta": paid or due,
                "recognizedRevenueDelta": paid or due,
                "deferredRevenueDelta": 0,
                "note": "订场收入已明确，小程序收款",
            }
        )
        return result

    if "未扣课包" in answer and "微信单独支付" in answer:
        result.update(
            {
                "parseType": "cash_course_income",
                "businessType": "课程收入",
                "paymentChannel": "微信",
                "recognizeRevenue": "是",
                "cashDelta": paid or due,
                "recognizedRevenueDelta": paid or due,
                "deferredRevenueDelta": 0,
                "note": "明确不是课包划扣，而是微信单独支付",
            }
        )
        return result

    if "大众" in answer or answer.startswith("券") or "微信补" in answer:
        coupon = None
        for pattern in [r"大众点评?(\d+(?:\.\d+)?)", r"大众(\d+(?:\.\d+)?)", r"券(\d+(?:\.\d+)?)"]:
            matched = re.search(pattern, answer)
            if matched:
                coupon = float(matched.group(1))
                break
        wechat = None
        matched = re.search(r"微信补(\d+(?:\.\d+)?)", answer)
        if matched:
            wechat = float(matched.group(1))
        if coupon is None and wechat is not None and paid:
            coupon = max(paid - wechat, 0)
        if coupon is None and "课程收入 /大众点评" in answer and paid:
            coupon = paid
        if coupon is None:
            result["status"] = "pending"
            result["parseType"] = "dp_coupon"
            result["businessType"] = "大众点评券"
            result["paymentChannel"] = "大众点评+微信补差"
            result["recognizeRevenue"] = "是"
            result["cashDelta"] = paid
            result["recognizedRevenueDelta"] = paid
            result["deferredRevenueDelta"] = 0
            result["note"] = "没能识别券金额"
            return result
        if wechat is None:
            wechat = max(paid - coupon, 0)
        result.update(
            {
                "parseType": "dp_coupon",
                "businessType": "大众点评券",
                "paymentChannel": "大众点评+微信补差" if wechat else "大众点评",
                "recognizeRevenue": "是",
                "cashDelta": paid,
                "recognizedRevenueDelta": paid,
                "deferredRevenueDelta": 0,
                "couponAmount": coupon,
                "wechatTopupAmount": wechat,
                "note": "可按大众点评券拆分处理",
            }
        )
        return result

    if "扣" in answer and ("/" in answer or ("," in answer and ("课包" in answer or "训练营" in answer))):
        student_name = ""
        package_name = ""
        lesson_count = ""
        if "/" in answer:
            parts = answer.split("/")
            student_name = parts[0].strip()
            package_name = parts[1].strip() if len(parts) > 1 else ""
            matched = re.search(r"扣\s*(\d+(?:\.\d+)?)", answer)
            lesson_count = float(matched.group(1)) if matched else ""
        else:
            parts = [part.strip() for part in answer.split(",") if part.strip()]
            student_name = parts[0] if parts else ""
            package_name = parts[1] if len(parts) > 1 else ""
            matched = re.search(r"扣\s*(\d+(?:\.\d+)?)", answer)
            lesson_count = float(matched.group(1)) if matched else ""
        amount = due or paid
        result.update(
            {
                "parseType": "package_consume",
                "businessType": "课包划扣",
                "paymentChannel": "课包划扣",
                "recognizeRevenue": "是",
                "cashDelta": 0,
                "recognizedRevenueDelta": amount,
                "deferredRevenueDelta": -amount if amount else 0,
                "studentName": student_name,
                "packageName": package_name,
                "lessonCount": lesson_count,
                "note": "课包/训练营扣减已明确",
            }
        )
        return result

    if answer == "课程":
        result["status"] = "pending"
        result["note"] = "只写了课程，没写收款情况"
        return result

    if answer == "oliver":
        result["status"] = "pending"
        result["note"] = "只写了人名，没说明这笔如何入账"
        return result

    result["status"] = "pending"
    result["note"] = "没有命中稳定规则"
    return result


def to_markdown_table(rows: list[dict[str, object]]) -> list[str]:
    if not rows:
        return ["无"]
    header = "| 原表行号 | 客户 | 你的回答 | 备注 |", "|---:|---|---|---|"
    body = [
        f"| {row['原表行号']} | {row['客户']} | {str(row['你的回答']).replace('|', '/')} | {str(row['note']).replace('|', '/')} |"
        for row in rows
    ]
    return [*header, *body]


def write_csv(path: Path, rows: list[dict[str, object]], columns: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def write_xlsx(path: Path, rows: list[dict[str, object]], columns: list[str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "结果"
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="把人工确认后的历史收入 Excel 分成可导入批次和挂起清单。")
    parser.add_argument("--input", required=True, help="待确认 Excel 路径")
    parser.add_argument("--output-dir", required=True, help="输出目录")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(input_path)
    sheet = workbook["待你确认"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    header_index = {header: offset + 1 for offset, header in enumerate(headers) if header}

    rows = []
    for row_number in range(2, sheet.max_row + 1):
        row = {header: sheet.cell(row_number, header_index[header]).value for header in header_index}
        if sheet.max_column >= 15:
            row["补充备注"] = sheet.cell(row_number, 15).value
        normalized = classify(row)
        row["你的回答"] = "" if row.get("你的回答") is None else str(row.get("你的回答")).strip()
        rows.append({**row, **normalized})

    ready_rows = [row for row in rows if row["status"] == "ready"]
    split_rows = [row for row in rows if row["status"] == "split"]
    pending_rows = [row for row in rows if row["status"] == "pending"]
    missing_rows = [row for row in rows if row["status"] == "missing"]
    blocked_rows = [*split_rows, *pending_rows, *missing_rows]

    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    base_name = f"confirmed-income-batch-{timestamp}"
    csv_columns = [
        "原表行号",
        "日期时间",
        "客户",
        "收入类型",
        "支付方式",
        "应收",
        "实收",
        "备注",
        "你的回答",
        "补充备注",
        "parseType",
        "businessType",
        "campusId",
        "campusName",
        "accountingScope",
        "paymentChannel",
        "recognizeRevenue",
        "cashDelta",
        "recognizedRevenueDelta",
        "deferredRevenueDelta",
        "studentName",
        "packageName",
        "lessonCount",
        "couponAmount",
        "wechatTopupAmount",
        "note",
    ]
    pending_columns = [
        "原表行号",
        "日期时间",
        "客户",
        "收入类型",
        "支付方式",
        "应收",
        "实收",
        "备注",
        "你的回答",
        "补充备注",
        "status",
        "note",
    ]

    ready_csv = output_dir / f"{base_name}-ready.csv"
    ready_xlsx = output_dir / f"{base_name}-ready.xlsx"
    blocked_xlsx = output_dir / f"{base_name}-blocked.xlsx"
    report_md = output_dir / f"{base_name}-report.md"

    write_csv(ready_csv, ready_rows, csv_columns)
    write_xlsx(ready_xlsx, ready_rows, csv_columns)
    write_xlsx(blocked_xlsx, blocked_rows, pending_columns)

    summary = Counter(row["parseType"] for row in ready_rows)
    lines = [
        f"# 人工确认历史收入批次报告",
        "",
        f"- 生成时间：{datetime.now().isoformat(timespec='seconds')}",
        f"- 输入文件：`{input_path}`",
        f"- 总行数：{len(rows)}",
        f"- 可直接导入：{len(ready_rows)}",
        f"- 仍需挂起：{len(blocked_rows)}",
        f"- 需要拆分：{len(split_rows)}",
        f"- 仍待确认：{len(pending_rows)}",
        f"- 回答缺失：{len(missing_rows)}",
        "",
        "## 可导入类型分布",
        "",
    ]
    for parse_type, count in summary.most_common():
        lines.append(f"- `{parse_type}`：{count}")
    lines.extend(["", "## 挂起清单", ""])
    lines.extend(to_markdown_table(blocked_rows))
    report_md.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"ready_csv={ready_csv}")
    print(f"ready_xlsx={ready_xlsx}")
    print(f"blocked_xlsx={blocked_xlsx}")
    print(f"report_md={report_md}")
    print(
        f"summary total={len(rows)} ready={len(ready_rows)} split={len(split_rows)} pending={len(pending_rows)} missing={len(missing_rows)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
